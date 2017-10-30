import time
import openpyxl
import urllib.request
import pycurl
from io import BytesIO
from bs4 import BeautifulSoup

class OpenRice(object):

    def get_url_data(self, district_id, page):
        data = {}
        data['districtId'] = district_id
        data['page'] = page
        return data

    def get_html(self, url):
        buffer = BytesIO()

        for i in range(0, 5):
            curl = pycurl.Curl()
            curl.setopt(pycurl.URL, url)
            curl.setopt(pycurl.USERAGENT, 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/602.2.14 (KHTML, like Gecko) Version/10.0.1 Safari/602.2.14')
            curl.setopt(pycurl.CONNECTTIMEOUT, 30)
            curl.setopt(pycurl.TIMEOUT, 300)
            curl.setopt(pycurl.WRITEDATA, buffer)

            try:
                curl.perform()
                curl.close()
                break
            except:
                curl.close()

                print('')
                print('# Exception!', ', retry: ', str(i + 1))
                print('')

                time.sleep(5)

        html = buffer.getvalue()

        return html.decode('utf-8')

    def find_restaurant_url(self, district):
        results = list()

        for page in range(1, 18):
            data = self.get_url_data(district, page)
            url_values = urllib.parse.urlencode(data)

            url = 'http://www.openrice.com/zh/macau/restaurants'
            full_url = url + '?' + url_values

            html = self.get_html(full_url)
            soup = BeautifulSoup(html, 'lxml')

            lis = soup.find('ul', 'sr1-listing-content-cells pois-restaurant-list js-poi-list-content-cell-container').find_all('li')

            for li in lis:
                h2 = li.find('h2', 'title-name')

                if h2 is None:
                    continue

                link = h2.find('a').get('href')

                results.append({'poi_id': li.find('section').get('data-poi-id'), 'url': urllib.parse.urljoin(url, link)})

            section = soup.find('section', 'js-pois-pagination pull-right')

            if section.find('div', 'or-sprite common_pagination_more_r_desktop') is None or page is 17:
                print('page:', page)
                break

        return results

    def main_find_restaurant_url(self):
        file_name = 'MC_OpenRice'
        district_sheet_name = 'district'
        restaurant_sheet_name = 'restaurant'

        start_time = time.time()
        print('# Load Excel...')

        wb = openpyxl.load_workbook(file_name + '.xlsx', read_only = False)
        district_sheet = wb.get_sheet_by_name(district_sheet_name)
        restaurant_sheet = wb.get_sheet_by_name(restaurant_sheet_name)

        loading_time = time.time()

        print('# Query...')
        print('# Total Query Data:', district_sheet.max_row - 1)
        print('')

        for rowNum in range(2, district_sheet.max_row + 1):
            district_name_cell = district_sheet.cell(row = rowNum, column = 2)
            district_id_cell = district_sheet.cell(row = rowNum, column = 3)

            print(rowNum - 1, district_name_cell.value)

            results = self.find_restaurant_url(district_id_cell.value)

            print('count:', len(results))
            print('')

            for result in results:
                restaurant_row_num = restaurant_sheet.max_row + 1

                poi_id_cell = restaurant_sheet.cell(row = restaurant_row_num, column = 1)
                url_cell = restaurant_sheet.cell(row = restaurant_row_num, column = 2)

                poi_id_cell.value = result['poi_id']
                url_cell.value = result['url']

        query_time = time.time()

        print('# Save Excel...')
        print('# Total Save Data:', restaurant_sheet.max_row - 1)

        wb.save(file_name + '.xlsx')

        save_time = time.time()

        print('')
        print('# Loadind Time:', loading_time - start_time, 's')
        print('# Query Time:', query_time - loading_time, 's')
        print('# Save Time:', save_time - query_time, 's')
        print('# Total Time:', save_time - start_time, 's')

    def find_restaurant_data(self, url):
        html = self.get_html(url)

        soup = BeautifulSoup(html, 'lxml')

        result = {}

        if soup.find('span', 'poi-with-other-status') is not None:
            return result

        result['name_c'] = soup.find('span', 'name').getText().strip()

        name_e = soup.find('div', 'smaller-font-name')
        result['name_e'] = name_e.getText().strip() if name_e is not None else ''

        result['district'] = soup.find('div', 'header-poi-district dot-separator').find('a').getText().strip()
        result['address_c'] = soup.find('div', 'address-info-section').find('div', 'content').find('a').getText().strip()
        result['address_e'] = soup.select('div .address-section > div')[1].find('a').getText().strip()

        phone = soup.find('section', 'telephone-section')
        result['phone'] = phone.find('div', 'content').getText().strip() if phone is not None else ''

        hours = soup.find('section', 'opening-hours-and-normal-notice-section')
        result['hours'] = hours.find('div', 'text').getText().strip() if hours is not None else ''

        result['price'] = soup.find('div', 'header-poi-price dot-separator').find('a').getText().strip()
        result['lng'] = soup.find('div', id = 'poi-report-map-container').get('data-longitude').strip()
        result['lat'] = soup.find('div', id  = 'poi-report-map-container').get('data-latitude').strip()

        return result

    def main_find_restaurant_data(self):
        file_name = 'MC_OpenRice'
        sheet_name = 'restaurant'

        start_time = time.time()
        print('# Load Excel...')

        wb = openpyxl.load_workbook(file_name + '.xlsx', read_only = False)
        sheet = wb.get_sheet_by_name(sheet_name)

        loading_time = time.time()

        print('# Query...')
        print('# Total Data:', sheet.max_row - 1)
        print('')

        for rowNum in range(2, sheet.max_row + 1):
            url_cell = sheet.cell(row = rowNum, column = 2)
            name_c_cell = sheet.cell(row = rowNum, column = 3)
            name_e_cell = sheet.cell(row = rowNum, column = 4)
            district_cell = sheet.cell(row = rowNum, column = 5)
            address_c_cell = sheet.cell(row = rowNum, column = 6)
            address_e_cell = sheet.cell(row = rowNum, column = 7)
            phone_cell = sheet.cell(row = rowNum, column = 8)
            hours_cell = sheet.cell(row = rowNum, column = 10)
            price_cell = sheet.cell(row = rowNum, column = 11)
            lng_cell = sheet.cell(row = rowNum, column = 12)
            lat_cell = sheet.cell(row = rowNum, column = 13)

            result = self.find_restaurant_data(url_cell.value)

            if len(result) == 0:
                continue

            name_c_cell.value = result['name_c']
            name_e_cell.value = result['name_e']
            district_cell.value = result['district']
            address_c_cell.value = result['address_c']
            address_e_cell.value = result['address_e']
            phone_cell.value = result['phone']
            hours_cell.value = result['hours']
            price_cell.value = result['price']
            lng_cell.value = result['lng']
            lat_cell.value = result['lat']

            print(str(rowNum - 1), name_c_cell.value)

        query_time = time.time()

        print('')
        print('# Save Excel...')

        wb.save(file_name + '.xlsx')

        save_time = time.time()

        print('')
        print('# Loadind Time:', loading_time - start_time, 's')
        print('# Query Time:', query_time - loading_time, 's')
        print('# Save Time:', save_time - query_time, 's')
        print('# Total Time:', save_time - start_time, 's')

if __name__ == '__main__':
    # OpenRice().main_find_restaurant_url()
    OpenRice().main_find_restaurant_data()
